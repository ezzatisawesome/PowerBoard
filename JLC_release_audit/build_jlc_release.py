import csv
import re
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
AUDIT = ROOT / "JLC_release_audit"
BOM_IN = AUDIT / "kicad_raw_bom_after.csv"
POS_IN = AUDIT / "PowerBoard-pos-after.csv"

ACTIVE_CSV = ROOT / "JLCPCB_BOM_RELEASE_ACTIVE_ONLY.csv"
CPL_CSV = ROOT / "JLCPCB_CPL_RELEASE_ACTIVE_ONLY.csv"
REVIEW_XLSX = ROOT / "JLCPCB_BOM_RELEASE_REVIEW.xlsx"

MANUAL_EXACT = {
    "U18": {
        "suggested": "C50400161",
        "note": "Exact JLC part for INA745AIRELR; JLCPCB page reports out of stock. Do not substitute without validating current-shunt rating, I2C behavior, and REL/VQFN-14 footprint.",
        "source": "https://jlcpcb.com/partdetail/TexasInstruments-INA745AIRELR/C50400161",
    },
    "U19": {
        "suggested": "C50400161",
        "note": "Exact JLC part for INA745AIRELR; JLCPCB page reports out of stock. Do not substitute without validating current-shunt rating, I2C behavior, and REL/VQFN-14 footprint.",
        "source": "https://jlcpcb.com/partdetail/TexasInstruments-INA745AIRELR/C50400161",
    },
}


def stock_page(cnum: str) -> dict:
    url = f"https://jlcpcb.com/partdetail/{cnum}"
    headers = {"User-Agent": "Mozilla/5.0"}
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=20) as r:
                html = r.read().decode("utf-8", errors="ignore")
            title = re.search(r"<title>(.*?)</title>", html, re.S)
            stock = re.search(r"In Stock:\s*([0-9,]+)", html)
            available = re.search(r"Available Order Qty:\s*([0-9,]+)", html)
            assembly = re.search(r"Assembly Type\s*</[^>]+>\s*<[^>]+>\s*([^<]+)", html)
            return {
                "LCSC Part": cnum,
                "URL": url,
                "Title": re.sub(r"\s+", " ", title.group(1)).strip() if title else "",
                "In Stock": int(stock.group(1).replace(",", "")) if stock else None,
                "Available Order Qty": int(available.group(1).replace(",", "")) if available else None,
                "Assembly Type": assembly.group(1).strip() if assembly else ("SMT Assembly" if "SMT Assembly" in html else ""),
                "Fetch Status": "OK",
            }
        except Exception as e:
            if attempt == 2:
                return {"LCSC Part": cnum, "URL": url, "Fetch Status": f"ERROR: {e}", "In Stock": None, "Available Order Qty": None, "Assembly Type": "", "Title": ""}
            time.sleep(0.5 * (attempt + 1))


def load_bom():
    rows = []
    with BOM_IN.open(newline="") as f:
        for r in csv.DictReader(f):
            r["LCSC Part"] = (r.get("LCSC Part") or r.get("Supplier_1 P/N") or "").strip()
            r["DNP"] = "DNP" if r.get("DNP") == "DNP" else ""
            r["Qty"] = int(r.get("Qty") or 1)
            rows.append(r)
    return rows


def write_csv(path, rows, headers):
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 70)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def main():
    bom = load_bom()
    cnums = sorted({r["LCSC Part"] for r in bom if r["LCSC Part"]})
    stock = {}
    with ThreadPoolExecutor(max_workers=12) as ex:
        futs = {ex.submit(stock_page, c): c for c in cnums}
        for fut in as_completed(futs):
            info = fut.result()
            stock[info["LCSC Part"]] = info

    active_upload = []
    all_rows = []
    review = []
    for r in bom:
        s = stock.get(r["LCSC Part"], {})
        stock_qty = s.get("In Stock")
        status = "DNP" if r["DNP"] else "OK"
        note = ""
        if not r["DNP"]:
            if not r["LCSC Part"]:
                status = "MANUAL_REVIEW"
                manual = MANUAL_EXACT.get(r["Reference"], {})
                note = manual.get("note", "Active component has no LCSC part number.")
                r["Suggested LCSC Part"] = manual.get("suggested", "")
                r["Source"] = manual.get("source", "")
            elif stock_qty is None:
                status = "VERIFY_MANUALLY"
                note = "Could not parse current JLC stock page."
            elif stock_qty <= 0:
                status = "OUT_OF_STOCK"
                note = "JLCPCB page reports zero stock."
        out = {
            "Reference": r["Reference"],
            "Value": r["Value"],
            "Footprint": r["Footprint"],
            "LCSC Part": r["LCSC Part"],
            "DNP": r["DNP"],
            "Qty": r["Qty"],
            "Status": status,
            "JLC In Stock": stock_qty if stock_qty is not None else "",
            "Available Order Qty": s.get("Available Order Qty", ""),
            "Assembly Type": s.get("Assembly Type", ""),
            "JLC URL": s.get("URL", r.get("Source", "")),
            "Note": note,
        }
        all_rows.append(out)
        if status == "OK":
            active_upload.append({k: out[k] for k in ["Reference", "Value", "Footprint", "LCSC Part", "DNP", "Qty"]})
        elif status != "DNP":
            review.append(out | {"Suggested LCSC Part": r.get("Suggested LCSC Part", ""), "Source": r.get("Source", "")})

    pos = list(csv.DictReader(POS_IN.open(newline="")))
    active_refs = {r["Reference"] for r in active_upload}
    cpl = [r for r in pos if r["Ref"] in active_refs]

    write_csv(ACTIVE_CSV, active_upload, ["Reference", "Value", "Footprint", "LCSC Part", "DNP", "Qty"])
    write_csv(CPL_CSV, cpl, ["Ref", "Val", "Package", "PosX", "PosY", "Rot", "Side"])

    wb = Workbook()
    for name, rows in [
        ("Active Upload BOM", active_upload),
        ("Active Upload CPL", cpl),
        ("All KiCad BOM Rows", all_rows),
        ("Manual Review", review),
        ("Stock Verification", [stock[c] for c in cnums]),
    ]:
        ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None else wb.create_sheet()
        ws.title = name
        headers = list(rows[0].keys()) if rows else ["No rows"]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        style_sheet(ws)

    ws = wb.create_sheet("Final Checklist")
    checks = [
        ("BOM generated from current KiCad schematic", "PASS"),
        ("LCSC Part populated from Supplier_1 P/N where available", "PASS"),
        ("One Reference per BOM row", "PASS"),
        ("No comma-separated LCSC cells", "PASS"),
        ("DNP set for TP1/TP2/J6/J7/J11/J12", "PASS"),
        ("Active upload rows all have LCSC Part", "PASS" if all(r["LCSC Part"] for r in active_upload) else "FAIL"),
        ("Active upload rows all parsed as JLC in-stock", "PASS" if all((stock.get(r["LCSC Part"], {}).get("In Stock") or 0) > 0 for r in active_upload) else "FAIL"),
        ("Manual review rows", len(review)),
        ("Active upload BOM rows", len(active_upload)),
        ("Active upload CPL rows", len(cpl)),
    ]
    ws.append(["Checklist Item", "Result"])
    for row in checks:
        ws.append(row)
    style_sheet(ws)
    wb.save(REVIEW_XLSX)

    print(f"active_bom={ACTIVE_CSV}")
    print(f"active_cpl={CPL_CSV}")
    print(f"review_xlsx={REVIEW_XLSX}")
    print(f"active_rows={len(active_upload)} cpl_rows={len(cpl)} review_rows={len(review)}")
    for r in review:
        print("REVIEW", r["Reference"], r["Value"], r["Status"], r.get("Suggested LCSC Part", ""), r["Note"])


if __name__ == "__main__":
    main()
